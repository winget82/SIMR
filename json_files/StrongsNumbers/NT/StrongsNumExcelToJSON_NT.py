from openpyxl import load_workbook
strongskjvot_nmbrlst = []
strkjvot = load_workbook('P:/python_work/nates/SIMR/reference_files/StrongsNumbers/NT/READY/STRONGSNMBRS_KJV_TITUS.xlsx')#UPDATE THIS FOR EACH BOOK
sheet = strkjvot.active

#READ EACH ROW OF CELL(S) IN FROM EXCEL SHEET AND APPEND TO LIST ADDING SPACE AT END
for row in sheet.iter_rows(min_row=1, min_col=1, max_row=151, max_col=1):#UPDATE THIS FOR EACH BOOK
	for cell in row:
		strongskjvot_nmbrlst.append(str(cell.value) + ' ')

jointlst = ''.join(strongskjvot_nmbrlst)

# now use regex
import re

#"Split string by the occurrences of pattern. If capturing parentheses
#are used in pattern, then the text of all groups in the pattern are
#also returned as part of the resulting list."

#USE REGEX TO SPLIT ON PATTERN, RETAIN THE PATTERN, AND DELETE THE FIRST LIST ENTRY
splitter = re.compile("(\s*\d+:\d+)\s*")
scripture = re.split(splitter, jointlst)
scripture.pop(0)

#TURN INTO LIST OF LISTS [REFERENCE,VERSE]
script2 = []
#print(scripture)
for i in range(0,len(scripture), 2):
	x = str(scripture[i]), str(scripture[i+1])
	script2.append(list(x))

#INSERT BIBLE BOOK TO REFERENCES i[0] of each list
book = 'Titus'#UPDATE THIS FOR EACH BOOK

#Add book as index 0 in each nested list
for i in script2:
	i.insert(0, book)
	i.insert(0, '|')
	i.insert(3, '+')
	new = [j for p in script2 for j in p]
new2 = ''.join(new)
final_book = [sverse.split('+') for sverse in new2.split('|') if sverse.split('+')]
del final_book[0]
print(final_book)

#SEE P.150FF IN AUTOMATE THE BORING STUFF WITH PYTHON
#SEE ALSO - https://docs.microsoft.com/en-us/dotnet/standard/base-types/regular-expression-language-quick-reference


#EXPORT SCRIPT2 TO JSON FILE TO UTILIZE IN SIMR
import json


#WRITE TO JSON FILE
filename = 'Titus_json.json'
with open(filename, 'w') as f_obj:
	json.dump(final_book, f_obj)
